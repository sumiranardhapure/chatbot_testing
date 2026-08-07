#!/usr/bin/env python3
"""
Golden Dataset Phase 1 Rebuild
- Keep only 'Standard' input type (1 row per scenario)
- Keep only paddy/rice rows
- Keep only monsoon onset scenarios
- Trim to essential columns (including wrong alert cols)
- Apply clean formatting: text wrap, column widths, frozen header
Output: golden_dataset_v2.xlsx
"""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent.parent.parent
DATA_DIR    = BASE_DIR / "data"
INPUT_FILE  = DATA_DIR / "golden_dataset.xlsx"
OUTPUT_FILE = DATA_DIR / "golden_dataset_v2.xlsx"

MONSOON_KEYWORDS = ["delayed", "onset", "early season drought", "normal"]

KEEP_COLUMNS = [
    "District",
    "Specific Scenario",
    "Land Type",
    "Irrigation Available",
    "Crop / Animal",
    "Expected: Seed Varieties",
    "Expected: Farming Practices",
    "Expected: Fertilizer Dose",
    "Expected: Chemicals",
    "Expected: Infrastructure",
    "All Expected Keywords",
    "Wrong Variety Alert (A)",
    "Wrong Land Type Alert (B)",
    "Wrong District Alert (C)",
    "Source Section",
]

# Column widths (characters)
COL_WIDTHS = {
    "District":                  14,
    "Specific Scenario":         32,
    "Land Type":                 20,
    "Irrigation Available":      14,
    "Crop / Animal":             14,
    "Expected: Seed Varieties":  38,
    "Expected: Farming Practices": 42,
    "Expected: Fertilizer Dose": 32,
    "Expected: Chemicals":       42,
    "Expected: Infrastructure":  32,
    "All Expected Keywords":     55,
    "Wrong Variety Alert (A)":   38,
    "Wrong Land Type Alert (B)": 38,
    "Wrong District Alert (C)":  38,
    "Source Section":            18,
}

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(size=9)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER  = Border(
    bottom=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="E8E8E8"),
)
ALT_FILL = PatternFill("solid", fgColor="F5F8FC")


def is_monsoon_scenario(val):
    if pd.isna(val):
        return False
    return any(kw in str(val).lower() for kw in MONSOON_KEYWORDS)


def is_paddy_row(val):
    if pd.isna(val):
        return False
    return bool(pd.Series([val]).str.lower().str.contains("rice|paddy", na=False).iloc[0])

BULLET_COLS = [
    "Expected: Seed Varieties",
    "Expected: Farming Practices",
    "Expected: Fertilizer Dose",
    "Expected: Chemicals",
    "Expected: Infrastructure",
    "All Expected Keywords",
]

def to_bullets(val):
    if pd.isna(val) or str(val).strip() == "":
        return val
    items = [item.strip() for item in str(val).split(";") if item.strip()]
    return "\n".join(f"• {item}" for item in items)

def process_sheet(df, sheet_name):
    df.columns = [str(c).strip() for c in df.columns]
    original_rows = len(df)
    notes = []

    # 1. Keep only Standard input type
    if "Input Type" in df.columns:
        has_standard = df["Input Type"].str.strip().str.lower().eq("standard").any()
        if has_standard:
            df = df[df["Input Type"].str.strip().str.lower() == "standard"]
        else:
            notes.append("WARNING: No 'Standard' input type — kept all rows")
    else:
        notes.append("WARNING: 'Input Type' column missing — kept all rows")
    after_input = len(df)

    # 2. Keep only paddy/rice rows
    if "Crop / Animal" in df.columns:
        df = df[df["Crop / Animal"].apply(is_paddy_row)]
    else:
        notes.append("WARNING: 'Crop / Animal' column missing — skipped paddy filter")
    after_paddy = len(df)

    # 3. Keep only monsoon onset scenarios
    if "Specific Scenario" in df.columns:
        df = df[df["Specific Scenario"].apply(is_monsoon_scenario)]
    else:
        notes.append("WARNING: 'Specific Scenario' column missing — skipped scenario filter")
    after_scenario = len(df)

    # 4. Trim to essential columns
    available = [c for c in KEEP_COLUMNS if c in df.columns]
    missing   = [c for c in KEEP_COLUMNS if c not in df.columns]
    df = df[available].reset_index(drop=True)

    if missing:
        notes.append(f"MISSING COLS: {missing}")

        # 5. Convert keyword columns to bullet points
    for col in BULLET_COLS:
        if col in df.columns:
            df[col] = df[col].apply(to_bullets)

    print(f"  {sheet_name:22s}: {original_rows:4d} → {after_input:4d} (input type)"
          f" → {after_paddy:4d} (paddy) → {after_scenario:4d} (monsoon onset)")
    for n in notes:
        print(f"    {n}")

    return df


def format_sheet(ws, col_names):
    """Apply formatting to a worksheet."""
    # Header row
    for col_idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = DATA_FONT
            cell.alignment = WRAP_ALIGN
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

    # Column widths
    for col_idx, col_name in enumerate(col_names, start=1):
        width = COL_WIDTHS.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto row height hint (openpyxl doesn't auto-calc, set a generous default)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60


def main():
    print(f"Reading {INPUT_FILE} ...")
    xls = pd.read_excel(INPUT_FILE, sheet_name=None)

    total_in, total_out = 0, 0
    sheet_data = {}  # name -> processed df

    for sheet_name, df in xls.items():
        if sheet_name.strip().upper() in ("README", "README_LEGEND", "LEGEND"):
            sheet_data[sheet_name] = ("readme", df)
            print(f"  {'README':22s}: copied as-is")
            continue
        processed = process_sheet(df.copy(), sheet_name)
        total_in  += len(df)
        total_out += len(processed)
        sheet_data[sheet_name] = ("district", processed)

    # Write to Excel
    writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")
    for sheet_name, (kind, df) in sheet_data.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()

    # Apply formatting via openpyxl
    wb = load_workbook(OUTPUT_FILE)
    for sheet_name, (kind, df) in sheet_data.items():
        ws = wb[sheet_name]
        if kind == "readme":
            ws.freeze_panes = "A2"
            continue
        col_names = list(df.columns)
        format_sheet(ws, col_names)
    wb.save(OUTPUT_FILE)

    print(f"\nTotal: {total_in} rows → {total_out} rows kept")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
