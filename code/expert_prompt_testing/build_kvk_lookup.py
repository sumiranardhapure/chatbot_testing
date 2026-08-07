#!/usr/bin/env python3
"""Build a cleaned KVK district lookup table for the 11 chatbot districts.

Reads data/KVK Master Locations.xls, matches each chatbot district to its KVK
record, adds 3 districts found via web research (Kondagaon, Koriya, Sukma),
generates Google Maps URLs, and outputs a clean Excel lookup table.

Usage:
    python3 code/expert_prompt_testing/build_kvk_lookup.py

Output:
    data/kvk_district_lookup.xlsx

Colour coding in output:
    Blue header  — column headers
    Yellow row   — address sourced from web research (no GPS coords)
    Red row      — no KVK found for this district
"""

from __future__ import annotations

import urllib.parse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent.parent.parent
DATA_DIR    = BASE_DIR / "data"
INPUT_FILE  = DATA_DIR / "KVK Master Locations.xls"
OUTPUT_FILE = DATA_DIR / "kvk_district_lookup.xlsx"

# The 11 chatbot districts (from KNOWN_DISTRICTS in score_chatbot.py)
CHATBOT_DISTRICTS = sorted([
    "Balrampur", "Bijapur", "Dantewada", "Jashpur", "Kanker",
    "Kondagaon", "Koriya", "Narayanpur", "Sukma", "Surajpur", "Surguja",
])

# Mapping: chatbot district → (district value in KVK file, substring to match in kvkname)
# Used to locate the correct row in KVK Master Locations.xls
# - Balrampur has no CG entry; its farmers are served by Surguja-II(Balrampur)
# - Surguja has 3 rows; the main KVK is the one without "II" in the name
KVK_FILE_MATCHES: dict[str, tuple[str, str | None]] = {
    "Balrampur":  ("Surguja",    "II"),
    "Bijapur":    ("Bijapur",    None),
    "Dantewada":  ("Dantewada",  None),
    "Jashpur":    ("Jashpur",    None),
    "Kanker":     ("Kanker",     None),
    "Narayanpur": ("Narayanpur", None),
    "Surguja":    ("Surguja",    "Surguja"),  # excludes the "II" row
}

# Districts whose KVK was not found in the source file but was located via web research.
# Addresses are taken from official ICAR/IGKV pages; GPS coordinates were not published.
WEB_SOURCED: list[dict] = [
    {
        "district":         "Kondagaon",
        "kvkname":          "Krishi Vigyan Kendra, Kondagaon",
        "address":          "Village Purvi Borgaon, NH-30, Block-Pharasgaon, "
                            "District-Kondagaon (C.G.) 494 229",
        "latitude":         None,
        "longitude":        None,
        "hostorganization": "Indira Gandhi Krishi Vishwavidyalaya (IGKV), Raipur",
        "notes":            "Address from igkvkvkkondagaoncg.in — not in KVK Master Locations.xls",
    },
    {
        "district":         "Koriya",
        "kvkname":          "Krishi Vigyan Kendra, Korea (Koriya)",
        "address":          "Village-Salka, Post-Mansukh, Block-Baikunthpur, "
                            "District-Korea (Koriya), Pin-497335",
        "latitude":         None,
        "longitude":        None,
        "hostorganization": "ICAR",
        "notes":            "Address from kvkkoreaigkv.org — not in KVK Master Locations.xls",
    },
    {
        "district":         "Sukma",
        "kvkname":          "Krishi Vigyan Kendra, Sukma",
        "address":          "Village Murtonda, District-Sukma, Chhattisgarh",
        "latitude":         None,
        "longitude":        None,
        "hostorganization": "Indira Gandhi Krishi Vishwavidyalaya (IGKV), Raipur",
        "notes":            "Address from ICAR Chhattisgarh directory — not in KVK Master Locations.xls",
    },
]

SURAJPUR_PLACEHOLDER: dict = {
    "district":         "Surajpur",
    "kvkname":          "",
    "address":          "",
    "latitude":         None,
    "longitude":        None,
    "maps_url":         "",
    "hostorganization": "",
    "notes":            "No dedicated KVK found. Surajpur was carved from Surguja in 2012 "
                        "and does not yet appear in the ICAR KVK directory. "
                        "Farmers may be directed to Surguja KVK.",
}


def maps_url(lat, lon, address: str) -> str:
    if pd.notna(lat) and pd.notna(lon):
        return f"https://www.google.com/maps?q={lat},{lon}"
    if address and address.strip():
        return (
            "https://www.google.com/maps/search/?api=1&query="
            + urllib.parse.quote(address.strip())
        )
    return ""


def find_row(df_cg: pd.DataFrame, district_val: str, name_frag: str | None) -> pd.Series | None:
    """Return the best-matching CG KVK row for a chatbot district."""
    pool = df_cg[df_cg["district"].str.strip().str.lower() == district_val.lower()]
    if pool.empty:
        return None
    if name_frag:
        filtered = pool[pool["kvkname"].str.contains(name_frag, case=False, na=False)]
        if not filtered.empty:
            pool = filtered
    # Prefer rows that have GPS coordinates
    with_coords = pool.dropna(subset=["latitude", "longitude"])
    return with_coords.iloc[0] if not with_coords.empty else pool.iloc[0]


def main():
    print(f"Loading {INPUT_FILE.name} ...")
    raw = pd.read_excel(INPUT_FILE)
    raw.columns = [str(c).strip() for c in raw.columns]

    df_cg = raw[raw["statename"].str.strip().str.lower() == "chhattisgarh"].copy()
    print(f"  Chhattisgarh rows in source file: {len(df_cg)}")

    records: list[dict] = []

    for district in CHATBOT_DISTRICTS:
        if district in KVK_FILE_MATCHES:
            dist_val, name_frag = KVK_FILE_MATCHES[district]
            row = find_row(df_cg, dist_val, name_frag)
            if row is None:
                print(f"  WARNING: {district} — no match found in source file")
                continue
            lat  = row["latitude"]  if pd.notna(row.get("latitude"))  else None
            lon  = row["longitude"] if pd.notna(row.get("longitude")) else None
            addr = str(row.get("address") or row.get("address_og") or "").strip()
            note = "Source: KVK Master Locations.xls"
            if dist_val != district:
                note += f" (listed under district='{dist_val}', kvkname='{row['kvkname']}')"
            rec = {
                "district":         district,
                "kvkname":          str(row.get("kvkname", "")).strip(),
                "address":          addr,
                "latitude":         lat,
                "longitude":        lon,
                "maps_url":         maps_url(lat, lon, addr),
                "hostorganization": str(row.get("hostorganization", "")).strip(),
                "notes":            note,
            }
            records.append(rec)
            print(f"  {district:<15s} ✓  {rec['kvkname']}")

        elif district == "Surajpur":
            records.append(SURAJPUR_PLACEHOLDER)
            print(f"  {district:<15s} ✗  not found — placeholder added")

        else:
            entry = next((e for e in WEB_SOURCED if e["district"] == district), None)
            if entry:
                rec = dict(entry)
                rec["maps_url"] = maps_url(rec["latitude"], rec["longitude"], rec["address"])
                records.append(rec)
                print(f"  {district:<15s} ~  {rec['kvkname']} (web-sourced, no GPS)")

    df_out = pd.DataFrame(records, columns=[
        "district", "kvkname", "address",
        "latitude", "longitude", "maps_url",
        "hostorganization", "notes",
    ])

    # ── Write Excel ───────────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="KVK Lookup")
        ws = writer.sheets["KVK Lookup"]

        FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
        FILL_MISSING = PatternFill("solid", fgColor="FFC7CE")   # red  — no KVK found
        FILL_WEB     = PatternFill("solid", fgColor="FFEB9C")   # yellow — web-sourced

        for cell in ws[1]:
            cell.fill      = FILL_HEADER
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        NOTES_COL = 8   # column H (1-indexed)
        for row in ws.iter_rows(min_row=2):
            notes_val = str(row[NOTES_COL - 1].value or "")
            if "No dedicated KVK" in notes_val:
                fill = FILL_MISSING
            elif "web-sourced" in notes_val.lower() or "not in KVK Master" in notes_val:
                fill = FILL_WEB
            else:
                fill = None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        col_widths = {
            "A": 14,  # district
            "B": 48,  # kvkname
            "C": 58,  # address
            "D": 11,  # latitude
            "E": 11,  # longitude
            "F": 72,  # maps_url
            "G": 42,  # hostorganization
            "H": 60,  # notes
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"

    print(f"\nDone. {len(df_out)} districts written to {OUTPUT_FILE.name}")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
