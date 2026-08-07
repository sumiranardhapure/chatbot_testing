#!/usr/bin/env python3
"""Bulk runner — sends every prompt in Expert Prompt Testing.xlsx to the chatbot
and stores the English and Hindi responses in a new output Excel file.

Each row produces two independent API sessions (one per language preset) so both
responses can be collected for comparison.

Usage:
    python3 code/expert_prompt_testing/run_expert_prompts.py

Input:  data/Expert testing_prompt sheet_v2.xlsx
Output: output/prompt_testing/Expert_Prompt_Testing_Responses_v2.xlsx
"""

from __future__ import annotations

import base64
import json
import urllib.error
import urllib.request
import uuid
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

# ─── PATHS ────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent.parent
DATA_DIR   = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

INPUT_FILE  = DATA_DIR / "Expert testing_prompt sheet_v2.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "prompt_testing" / "Expert_Prompt_Testing_Responses_v2.xlsx"

# ─── API CONSTANTS ────────────────────────────────────────────────────────────
BASE_URL        = "https://wh-2e931fd08cae45d4a7f56236e5f69780.ecs.ap-south-1.on.aws/"
USERNAME        = "pxd-dil"
PASSWORD        = "p@ddyADVICE2theW0rld"
TIMEOUT_SECONDS = 120

# ─── SCENARIO MAP ─────────────────────────────────────────────────────────────
SCENARIO_MAP = {
    "early onset":             "early_onset",
    "normal onset":            "normal_onset",
    "delayed onset – 2 weeks": "delayed_onset_2w",
    "delayed onset - 2 weeks": "delayed_onset_2w",
    "delayed onset – 4 weeks": "delayed_onset_4w",
    "delayed onset - 4 weeks": "delayed_onset_4w",
    "delayed onset – 6 weeks": "delayed_onset_6w",
    "delayed onset - 6 weeks": "delayed_onset_6w",
    "delayed onset – 8 weeks": "delayed_onset_8w",
    "delayed onset - 8 weeks": "delayed_onset_8w",
}


# ─── HTTP HELPERS (same pattern as api_example.py) ───────────────────────────

def _auth_header() -> str:
    token = base64.b64encode(f"{USERNAME}:{PASSWORD}".encode("utf-8")).decode("ascii")
    return f"Basic {token}"


def post_json(path: str, payload: dict) -> dict:
    request = urllib.request.Request(
        f"{BASE_URL.rstrip('/')}{path}",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": _auth_header(),
        },
    )
    with urllib.request.urlopen(request, timeout=TIMEOUT_SECONDS) as response:
        return json.loads(response.read().decode("utf-8"))


# ─── SESSION RUNNER ───────────────────────────────────────────────────────────

def run_session(prompt_key: str, scenario_api_key: str | None, message_text: str) -> tuple[str, str]:
    """Run one complete session and return (reply, session_id).

    Raises urllib.error.HTTPError / URLError on failure.
    """
    session_id = f"expert-{prompt_key}-{uuid.uuid4()}"

    post_json("/prompt", {"session_id": session_id, "prompt_key": prompt_key})

    if scenario_api_key:
        post_json("/scenario", {"session_id": session_id, "scenario": scenario_api_key})

    result = post_json("/message", {"session_id": session_id, "text": message_text})
    return result["reply"], session_id


# ─── EXCEL STYLE HELPERS ──────────────────────────────────────────────────────

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_EN_HDR = PatternFill("solid", fgColor="2E75B6")
FILL_HI_HDR = PatternFill("solid", fgColor="375623")


def style_sheet(ws, n_original_cols: int):
    # Appended column order: Response(EN), Response(HI), Session ID(EN), Session ID(HI), Notes
    for i, cell in enumerate(ws[1], start=1):
        if i <= n_original_cols:
            cell.fill = FILL_HEADER
        elif i in (n_original_cols + 1, n_original_cols + 3):
            cell.fill = FILL_EN_HDR  # Response (English), Session ID (English)
        elif i in (n_original_cols + 2, n_original_cols + 4):
            cell.fill = FILL_HI_HDR  # Response (Hindi), Session ID (Hindi)
        else:
            cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading {INPUT_FILE.name}...")
    sheets_raw = pd.read_excel(INPUT_FILE, sheet_name=None)
    sheets_raw = {name: df for name, df in sheets_raw.items()}

    results: dict[str, pd.DataFrame] = {}
    total_rows = sum(len(df) for df in sheets_raw.values())
    processed = 0

    for sheet_name, df in sheets_raw.items():
        df = df.copy()
        # Normalise column headers
        df.columns = [str(c).strip() for c in df.columns]

        # v2 input already has empty placeholder columns for the output fields —
        # drop them so the script can append them cleanly in the correct position.
        df.drop(columns=[c for c in ["Response (English)", "Response (Hindi)", "Notes"]
                         if c in df.columns], inplace=True)

        n_original = len(df.columns)

        df["Response (English)"]    = ""
        df["Response (Hindi)"]      = ""
        df["Session ID (English)"]  = ""
        df["Session ID (Hindi)"]    = ""
        df["Notes"]                 = ""

        n_rows = len(df)

        for idx in range(n_rows):
            processed += 1
            row = df.iloc[idx]

            # Resolve scenario
            scenario_raw = str(row.get("Scenario", "")).strip()
            scenario_api = SCENARIO_MAP.get(scenario_raw.lower())
            notes_parts: list[str] = []
            if scenario_raw and not scenario_api:
                notes_parts.append(f"Unknown scenario: '{scenario_raw}'")

            prompt_text = str(row.get("Prompt Text", "")).strip()
            if not prompt_text:
                df.at[df.index[idx], "Notes"] = "Empty prompt — skipped"
                print(f"[{sheet_name} {idx+1}/{n_rows}] SKIP  empty prompt")
                continue

            en_reply = en_sid = hi_reply = hi_sid = ""

            # ── English session ───────────────────────────────────────────────
            try:
                en_reply, en_sid = run_session("english", scenario_api, prompt_text)
                en_status = "EN ✓"
            except (urllib.error.HTTPError, urllib.error.URLError) as e:
                err = e.read().decode("utf-8", "replace") if isinstance(e, urllib.error.HTTPError) else str(e.reason)
                notes_parts.append(f"EN error: {err[:120]}")
                en_status = "EN ✗"

            # ── Hindi session ─────────────────────────────────────────────────
            try:
                hi_reply, hi_sid = run_session("hindi", scenario_api, prompt_text)
                hi_status = "HI ✓"
            except (urllib.error.HTTPError, urllib.error.URLError) as e:
                err = e.read().decode("utf-8", "replace") if isinstance(e, urllib.error.HTTPError) else str(e.reason)
                notes_parts.append(f"HI error: {err[:120]}")
                hi_status = "HI ✗"

            df.at[df.index[idx], "Response (English)"]   = en_reply
            df.at[df.index[idx], "Response (Hindi)"]     = hi_reply
            df.at[df.index[idx], "Session ID (English)"] = en_sid
            df.at[df.index[idx], "Session ID (Hindi)"]   = hi_sid
            df.at[df.index[idx], "Notes"]                = "  |  ".join(notes_parts)

            sid_preview = en_sid[:28] if en_sid else "—"
            print(f"[{sheet_name} {idx+1}/{n_rows}] {en_status}  {hi_status}  session_en={sid_preview}…")

        results[sheet_name] = df

    # ── Write output ──────────────────────────────────────────────────────────
    print(f"\nWriting {OUTPUT_FILE.name}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, df in results.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            n_original = len(df.columns) - 5  # we added 5 columns
            style_sheet(ws, n_original)

            # Column widths
            # A-I = 9 original cols, J-N = 5 appended cols
            ws.column_dimensions["A"].width = 5   # #
            ws.column_dimensions["B"].width = 12  # Language
            ws.column_dimensions["C"].width = 22  # Scenario
            ws.column_dimensions["D"].width = 14  # District
            ws.column_dimensions["E"].width = 16  # Taluka
            ws.column_dimensions["F"].width = 14  # Land Type
            ws.column_dimensions["G"].width = 22  # Irrigation Access
            ws.column_dimensions["H"].width = 55  # Prompt Text
            ws.column_dimensions["I"].width = 30  # Test Type
            ws.column_dimensions["J"].width = 60  # Response (English)
            ws.column_dimensions["K"].width = 60  # Response (Hindi)
            ws.column_dimensions["L"].width = 36  # Session ID (English)
            ws.column_dimensions["M"].width = 36  # Session ID (Hindi)
            ws.column_dimensions["N"].width = 40  # Notes

    print(f"\nDone. {processed} rows processed across {len(results)} sheets.")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
