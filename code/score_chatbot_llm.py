#!/usr/bin/env python3
"""
CRIDA DACP Chatbot Scoring Script — LLM-as-judge edition
=========================================================
Scores chatbot responses (col L "English response (normal onset)") against
golden dataset keywords using Claude Sonnet 4.6 as the judge.
Restricted to Balrampur, Surajpur, and Jashpur districts.

Setup (run once in terminal):
    pip3 install pandas openpyxl anthropic
    export ANTHROPIC_API_KEY=sk-...

Usage:
    Run from any directory:
        python3 code/score_chatbot_llm.py
    Input files are read from data/, output is written to output/.

Output:
    Chatbot_Scoring_Results_llm.xlsx — two sheets: Summary, Standard Scores.
"""

import re
import json
import os
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import anthropic

BASE_DIR   = Path(__file__).parent.parent
DATA_DIR   = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

# ─── ANTHROPIC CLIENT ─────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
if not ANTHROPIC_API_KEY:
    raise EnvironmentError(
        "ANTHROPIC_API_KEY environment variable is not set.\n"
        "Run: export ANTHROPIC_API_KEY=sk-..."
    )

_ANTHROPIC_CLIENT = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
LLM_MODEL = "claude-sonnet-4-6"

# ─── PROMPT TEMPLATE ─────────────────────────────────────────────────────────
KEYWORD_JUDGE_PROMPT = """\
You are evaluating whether an agricultural chatbot response covers expected keywords.

Chatbot response:
\"\"\"
{response}
\"\"\"

Expected keywords (may include English and Hindi agricultural terms):
{keywords_list}

For each keyword, decide if the chatbot response captures the same concept — even if worded differently, abbreviated, or expressed in a different language. Be flexible: partial mentions count as long as meaning is conveyed. Don't penalize vague phrases in keywords missed and don't double count keywords if mentioned twice.

Return ONLY a JSON object with two arrays:
{{
  "captured": ["keyword1", "keyword2", ...],
  "missed":   ["keyword3", ...]
}}

Use the exact keyword strings from the list above. Do not add any explanation outside the JSON.\
"""

# ─── FILE PATHS ───────────────────────────────────────────────────────────────
PROMPT_BANK_FILE    = DATA_DIR / "Chatbot_Query_Bank_Surajpur_Balrampur_Jashpur_v2.xlsx"
GOLDEN_DATASET_FILE = DATA_DIR / "golden_dataset_v2.xlsx"
OUTPUT_FILE         = OUTPUT_DIR / "Chatbot_Scoring_Results_llm.xlsx"


# ─── SUB-DISTRICT → DISTRICT MAPPING ─────────────────────────────────────────
SUBDISTRICT_TO_DISTRICT = {
    "sukma":          "Sukma",
    "bijapur":        "Bijapur",
    "bhopalpatnam":   "Bijapur",
    "dantewada":      "Dantewada",
    "geedam":         "Dantewada",
    "kuakonda":       "Dantewada",
    "kondagaon":      "Kondagaon",
    "odgil":          "Kondagaon",
    "koriya":         "Koriya",
    "sonhat":         "Koriya",
    "jashpur":        "Jashpur",
    "bagicha":        "Jashpur",
    "pathalgaon":     "Jashpur",
    "pharagaon":      "Jashpur",
    "surguja":        "Surguja",
    "ambikapur":      "Surguja",
    "lundra":         "Surguja",
    "lakhanpur":      "Surguja",
    "surajpur":       "Surajpur",
    "chhindgarh":     "Sukma",
    "sitapur":        "Surajpur",
    "balrampur":      "Balrampur",
    "ramanujnagar":   "Balrampur",
    "batouli":        "Balrampur",
    "narayanpur":     "Narayanpur",
    "orchaa":         "Narayanpur",
    "kanker":         "Kanker",
    "charma":         "Kanker",
    "bhanupratappur": "Kanker",
    "pakhanjur":      "Kanker",
}
KNOWN_DISTRICTS = list(set(SUBDISTRICT_TO_DISTRICT.values()))

def extract_explicit_districts(text):
    if not text:
        return None
    text_lower = str(text).lower()
    for d in KNOWN_DISTRICTS:
        if d.lower() in text_lower:
            return d
    return None

# ─── LAND TYPE SYNONYM MAP ────────────────────────────────────────────────────
LAND_TYPE_SYNONYMS = {
    "marhan":  "upland",
    "tikra":   "upland",
    "gabhar":  "lowland",
    "mal":     "midland",
    "bharri":  "upland",
    "bhata":   "upland",
    "kanhar":  "lowland",
    "matasi":  "midland",
    "upariya": "upland",
    "dorsa":   "midland",
}


# ─── MONSOON-RELEVANT SCENARIO FILTER ────────────────────────────────────────
MONSOON_SCENARIO_KEYWORDS = [
    "delayed",
    "onset",
    "early season drought",
    "normal",
]


# ─── ADVERSARIAL FLAGGING PHRASES ────────────────────────────────────────────
FLAGGING_PHRASES = [
    "does not exist", "not valid", "incorrect", "not found", "please clarify",
    "please provide", "cannot find", "outside scope", "not covered",
    "not applicable", "please specify", "unable to", "invalid",
    "i need more information", "clarify", "what is your land type",
    "what is your district", "please tell me", "out of scope",
    "not available", "not recognized", "which district", "which land type",
    "not a valid", "soil type", "does not match", "wrong", "not exist",
    "mismatch", "not supported", "not calibrated",
    "कृपया", "जानकारी दें", "जिला बताएं", "भूमि प्रकार", "मुझे बताएं",
    "सही नहीं", "मान्य नहीं", "उपलब्ध नहीं", "please tell", "tell me your", "bata", "batao",
    "kaun sa", "कौन सा"
]


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def has_devanagari(text):
    return bool(re.search(r'[ऀ-ॿ]', str(text)))

def parse_key_inputs(raw):
    if pd.isna(raw):
        return None, None, None
    s = str(raw).strip()
    if s.lower() in ("no key input", "", "nan"):
        return "NO_KEY", None, None
    s = s.replace("\n", ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]
    subdistrict = land_type = irrigation = None
    for part in parts:
        pl = part.lower()
        if "sub-district" in pl or "subdistrict" in pl:
            subdistrict = re.sub(r"\s*sub-?district\s*", "", pl, flags=re.IGNORECASE).strip()
        elif "no irrigation" in pl:
            irrigation = "No"
        elif "irrigation available" in pl:
            irrigation = "Yes"
        else:
            cleaned = re.sub(r"\s*\bland\b\s*$", "", pl, flags=re.IGNORECASE).strip()
            if cleaned:
                land_type = cleaned
    if land_type:
        land_type = next((c for s, c in LAND_TYPE_SYNONYMS.items() if s in land_type), land_type)
    return subdistrict, land_type, irrigation

def is_monsoon_scenario(scenario_text):
    if pd.isna(scenario_text):
        return False
    sl = str(scenario_text).lower()
    return any(kw in sl for kw in MONSOON_SCENARIO_KEYWORDS)

def find_golden_rows(gd_sheets, district, land_type, irrigation):
    if district not in gd_sheets:
        return pd.DataFrame(), f"Sheet '{district}' not found"
    sheet = gd_sheets[district].copy()
    sheet = sheet[sheet["Crop / Animal"].str.lower().str.contains("rice", na=False)]
    if irrigation == "No":
        sheet = sheet[sheet["Irrigation Available"].isin(["No"])]
    elif irrigation == "Yes":
        sheet = sheet[sheet["Irrigation Available"].isin(["Yes", "Yes/No"])]
    if land_type:
        mask = sheet["Land Type"].str.lower().str.contains(re.escape(land_type), na=False)
        if mask.sum() == 0:
            words = [w for w in land_type.split() if len(w) > 2]
            if words:
                mask = sheet["Land Type"].str.lower().apply(lambda x: any(w in x for w in words))
        sheet = sheet[mask]
    if "Specific Scenario" in sheet.columns:
        sheet = sheet[sheet["Specific Scenario"].apply(is_monsoon_scenario)]
    else:
        print(f"  WARNING: 'Specific Scenario' column missing in '{district}' — skipping scenario filter")
    if sheet.empty:
        return pd.DataFrame(), "No row matches (district + land type + irrigation + monsoon scenario)"
    return sheet, "OK"

def golden_rows_label(golden_rows, district):
    excel_rows = [str(i + 2) for i in golden_rows.index.tolist()]
    return f"{district} · rows {', '.join(excel_rows)}"

def build_union_keywords(golden_rows):
    seen = set()
    union = []
    for _, grow in golden_rows.iterrows():
        AGRONOMIC_COLS = [
            "Expected: Seed Varieties",
            "Expected: Farming Practices",
            "Expected: Fertilizer Dose",
            "Expected: Chemicals",
            "Expected: Infrastructure",
        ]
        kw_str = " ; ".join(str(grow[c]) for c in AGRONOMIC_COLS if c in grow.index and pd.notna(grow[c]))
        if pd.isna(kw_str):
            continue
        for raw in str(kw_str).replace(";", "\n").split("\n"):
            sub_items = [i.strip() for i in re.sub(r'^[•.\s]+', '', raw).strip().split(",") if i.strip()]
            new_kws = [kw for kw in sub_items if kw and kw.lower() not in seen]
            seen.update(kw.lower() for kw in new_kws)
            union.extend(new_kws)
    return union


def score_keywords(response, keywords):
    """Score keywords against a response using Claude Sonnet 4.6 as the judge.

    Claude receives the full response and keyword list, then returns a JSON object
    identifying which keywords are captured (even via paraphrasing or translation).
    """
    if not keywords or pd.isna(response):
        return None, [], []

    resp = str(response).strip()
    n = len(keywords)
    kw_list_str = "\n".join(f"- {kw}" for kw in keywords)

    prompt = KEYWORD_JUDGE_PROMPT.format(
        response=resp,
        keywords_list=kw_list_str,
    )

    try:
        message = _ANTHROPIC_CLIENT.messages.create(
            model=LLM_MODEL,
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text.strip()
    except anthropic.APIError as e:
        print(f"  API error: {e}")
        return None, [], list(keywords)

    # Strip optional ```json ... ``` fences Claude may add
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        result = json.loads(raw)
    except json.JSONDecodeError:
        print(f"  JSON parse error. Raw response:\n{raw[:300]}")
        return None, [], list(keywords)

    # Case-insensitive reconciliation: map Claude's strings back to original keywords
    captured_lower = {s.lower() for s in result.get("captured", [])}

    matched = []
    missed = []
    for kw in keywords:
        if kw.lower() in captured_lower:
            matched.append(f"{kw} [~llm]")
        else:
            missed.append(kw)

    pct = round(len(matched) / n * 100, 1)
    return pct, matched, missed


def score_adversarial(response):
    if pd.isna(response) or str(response).strip() == "":
        return "FAIL", "No response recorded"
    resp_lower = str(response).lower()
    if any(phrase in resp_lower for phrase in FLAGGING_PHRASES):
        return "PASS", "Chatbot flagged the issue or asked for clarification"
    return "FAIL", "Chatbot answered directly — did not flag the problem"


# ─── EXCEL STYLE HELPERS ──────────────────────────────────────────────────────

FILL_GREEN     = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW    = PatternFill("solid", fgColor="FFEB9C")
FILL_RED       = PatternFill("solid", fgColor="FFC7CE")
FILL_GREY      = PatternFill("solid", fgColor="D9D9D9")
FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")
FILL_AUDIT_HDR = PatternFill("solid", fgColor="2E75B6")

def style_header(ws, audit_col_idx=None):
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = FILL_AUDIT_HDR if (audit_col_idx and i == audit_col_idx) else FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def wrap_data_cells(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading files...")
    pb = pd.read_excel(PROMPT_BANK_FILE)
    pb.columns = [c.strip() for c in pb.columns]
    gd_raw = pd.read_excel(GOLDEN_DATASET_FILE, sheet_name=None)
    gd = {name: df.rename(columns=lambda c: str(c).strip()) for name, df in gd_raw.items()}
    pb = pb.dropna(subset=["Q#"]).reset_index(drop=True)
    print(f"  Prompt rows to score: {len(pb)}")

    std_results = []

    for _, row in pb.iterrows():
        sr         = int(row["Q#"])
        prompt     = str(row["English Query"])
        response   = row["English response (normal onset)"]
        district   = str(row["District"]).strip() if pd.notna(row["District"]) else ""
        local_term = str(row["Land Type (local term)"]).strip().lower() if pd.notna(row["Land Type (local term)"]) else ""
        land_type  = LAND_TYPE_SYNONYMS.get(local_term, local_term) if local_term else None
        irrigation = "Yes" if str(row["Irrigation"]).strip() == "Yes" else "No"

        rec = {
            "Q#"                  : sr,
            "Prompt"              : prompt,
            "Chatbot Response"    : str(response) if pd.notna(response) else "",
            "Key Inputs"          : "",
            "Keywords (Reference)": "",
            "Keywords Captured"   : "",
            "Keywords Missed"     : "",
            "Score (%)"           : "",
            "Golden Dataset Rows" : "",
            "Notes"               : "",
        }

        if not district:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "District not specified"
            std_results.append(rec)
            continue

        rec["Key Inputs"] = f"{district} / {land_type or 'land type not specified'} / {'No irrigation' if irrigation == 'No' else 'Irrigation available'}"

        golden_rows, status = find_golden_rows(gd, district, land_type, irrigation)
        if golden_rows.empty:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = f"No golden dataset match — {status}"
            std_results.append(rec)
            continue

        missing_input = "not specified" in rec["Key Inputs"].lower()
        asks_clarification = "?" in str(response) and any(
            p in str(response).lower() for p in [
                "land type", "land-type", "upland", "lowland", "midland",
                "district", "block", "tehsil", "subdistrict",
                "irrigation", "soil type", "zameen", "bhoomi",
                "जमीन", "जिला", "सिंचाई", "भूमि"
            ]
        )
        if missing_input and asks_clarification:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "Correct behaviour — chatbot asked for missing key input"
            std_results.append(rec)
            continue

        rec["Golden Dataset Rows"] = golden_rows_label(golden_rows, district)
        union_kw = build_union_keywords(golden_rows)
        rec["Keywords (Reference)"] = ("• " + "\n• ".join(union_kw)) if union_kw else ""

        pct, matched, missed = score_keywords(response, union_kw)
        if pct is None:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "Keywords column empty in golden dataset or API error"
        else:
            rec["Score (%)"]         = pct
            rec["Keywords Captured"] = ("• " + "\n• ".join(matched)) if matched else ""
            rec["Keywords Missed"]   = ("• " + "\n• ".join(missed)) if missed else ""
            rec["Notes"] = "Low score — review chatbot response" if pct < 30 else ("Partial — some key advice missing" if pct < 60 else "")
        std_results.append(rec)

    std_df = pd.DataFrame(std_results)
    scores  = pd.to_numeric(std_df["Score (%)"], errors="coerce").dropna()

    summary_df = pd.DataFrame({
        "Metric": [
            "Total prompts tested",
            "Standard prompts scored", "Standard prompts – no match / no key input",
            "Average keyword score (%)", "Score >= 60% (good)", "Score 30-59% (partial)",
            "Score < 30% (poor)",
        ],
        "Value": [
            len(std_df),
            len(scores), len(std_df) - len(scores),
            f"{scores.mean():.1f}%" if len(scores) else "–",
            int((scores >= 60).sum()) if len(scores) else 0,
            int(((scores >= 30) & (scores < 60)).sum()) if len(scores) else 0,
            int((scores < 30).sum()) if len(scores) else 0,
        ],
    })

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]
        style_header(ws)
        set_col_widths(ws, {"A": 50, "B": 20})
        wrap_data_cells(ws)

        std_df.to_excel(writer, index=False, sheet_name="Standard Scores")
        ws = writer.sheets["Standard Scores"]
        style_header(ws, audit_col_idx=9)
        set_col_widths(ws, {
            "A": 6,  "B": 50, "C": 55, "D": 35,
            "E": 55, "F": 55, "G": 55, "H": 10,
            "I": 28, "J": 40,
        })
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            score_cell = row[7]
            val = score_cell.value
            if val == "N/A":
                score_cell.fill = FILL_GREY
            else:
                try:
                    s = float(val)
                    score_cell.fill = FILL_GREEN if s >= 60 else (FILL_YELLOW if s >= 30 else FILL_RED)
                except (TypeError, ValueError):
                    pass
        wrap_data_cells(ws)

    print(f"\n{'='*50}\nRESULTS SUMMARY\n{'='*50}")
    print(f"Total : {len(std_df)} prompts, {len(scores)} scored")
    if len(scores):
        print(f"  Avg score : {scores.mean():.1f}%  |  >=60%: {int((scores>=60).sum())}  |  30-59%: {int(((scores>=30)&(scores<60)).sum())}  |  <30%: {int((scores<30).sum())}")
    print(f"\nOutput saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
