#!/usr/bin/env python3
"""
CRIDA DACP Chatbot Scoring Script — RRF edition
================================================
Scores chatbot responses (col L "English response (normal onset)") against
golden dataset keywords using Reciprocal Rank Fusion.
Restricted to Balrampur, Surajpur, and Jashpur districts.

Setup (run once in terminal):
    pip3 install pandas openpyxl rapidfuzz indic-transliteration

Usage:
    Run from any directory:
        python3 code/score_chatbot_rrf.py
    Input files are read from data/, output is written to output/.

Output:
    Chatbot_Scoring_Results_rrf.xlsx — two sheets: Summary, Standard Scores.
"""

import re
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR   = Path(__file__).parent.parent
DATA_DIR   = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

try:
    from rapidfuzz import fuzz
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False
    print("WARNING: rapidfuzz not installed. Run: pip3 install rapidfuzz")

try:
    from indic_transliteration import sanscript
    from indic_transliteration.sanscript import transliterate
    INDIC_AVAILABLE = True
except ImportError:
    INDIC_AVAILABLE = False

try:
    from sentence_transformers import SentenceTransformer
    from sentence_transformers.util import cos_sim
    _SEMANTIC_MODEL = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
    SEMANTIC_AVAILABLE = True
except ImportError:
    SEMANTIC_AVAILABLE = False
    print("INFO: sentence-transformers not installed — semantic matching disabled.")
    print("      Run: pip install sentence-transformers")


# ─── TUNABLE THRESHOLDS ───────────────────────────────────────────────────────
SEMANTIC_THRESHOLD  = 0.60   # fallback floor when only semantic is available
FUZZY_THRESHOLD     = 75     # fallback floor (0–100) when only fuzzy is available

# ─── RRF PARAMETERS ───────────────────────────────────────────────────────────
# Small k amplifies rank differences vs. the IR default of 60, which compresses
# too much for short keyword lists (typically 10–30 items).
#
# With k=5 and 2 methods, RRF score range:
#   rank 1 in both  → 1/6 + 1/6 ≈ 0.333   ← strong match
#   rank 3 in both  → 1/8 + 1/8 ≈ 0.250   ← threshold
#   rank 5 in both  → 1/10 + 1/10 = 0.200 ← miss
RRF_K               = 5
RRF_MATCH_THRESHOLD = 0.25

# ─── FILE PATHS ───────────────────────────────────────────────────────────────
PROMPT_BANK_FILE    = DATA_DIR / "Chatbot_Query_Bank_Surajpur_Balrampur_Jashpur_v2.xlsx"
GOLDEN_DATASET_FILE = DATA_DIR / "golden_dataset_v2.xlsx"
OUTPUT_FILE         = OUTPUT_DIR / "Chatbot_Scoring_Results_rrf.xlsx"


# ─── SUB-DISTRICT → DISTRICT MAPPING ─────────────────────────────────────────
SUBDISTRICT_TO_DISTRICT = {
    "sukma":          "Sukma",
    "bijapur":        "Bijapur",
    "bhopalpatnam":   "Bijapur",
    "dantewada":      "Dantewada",
    "geedam":         "Dantewada",
    "kuakonda":       "Dantewada",
    "kondagaon":      "Kondagaon",
    "odgil":          "Kondagaon",
    "koriya":         "Koriya",
    "sonhat":         "Koriya",
    "jashpur":        "Jashpur",
    "bagicha":        "Jashpur",
    "pathalgaon":     "Jashpur",
    "pharagaon":      "Jashpur",
    "surguja":        "Surguja",
    "ambikapur":      "Surguja",
    "lundra":         "Surguja",
    "lakhanpur":      "Surguja",
    "surajpur":       "Surajpur",
    "chhindgarh":     "Sukma",
    "sitapur":        "Surajpur",
    "balrampur":      "Balrampur",
    "ramanujnagar":   "Balrampur",
    "batouli":        "Balrampur",
    "narayanpur":     "Narayanpur",
    "orchaa":         "Narayanpur",
    "kanker":         "Kanker",
    "charma":         "Kanker",
    "bhanupratappur": "Kanker",
    "pakhanjur":      "Kanker",
}
KNOWN_DISTRICTS = list(set(SUBDISTRICT_TO_DISTRICT.values()))

def extract_explicit_districts(text):
    if not text:
        return None
    text_lower = str(text).lower()
    for d in KNOWN_DISTRICTS:
        if d.lower() in text_lower:
            return d
    return None

# ─── LAND TYPE SYNONYM MAP ────────────────────────────────────────────────────
LAND_TYPE_SYNONYMS = {
    "marhan":  "upland",
    "tikra":   "upland",
    "gabhar":  "lowland",
    "mal":     "midland",
    "bharri":  "upland",
    "bhata":   "upland",
    "kanhar":  "lowland",
    "matasi":  "midland",
    "upariya": "upland",
    "dorsa":   "midland",
}


# ─── MONSOON-RELEVANT SCENARIO FILTER ────────────────────────────────────────
MONSOON_SCENARIO_KEYWORDS = [
    "delayed",
    "onset",
    "early season drought",
    "normal",
]


# ─── ADVERSARIAL FLAGGING PHRASES ────────────────────────────────────────────
FLAGGING_PHRASES = [
    "does not exist", "not valid", "incorrect", "not found", "please clarify",
    "please provide", "cannot find", "outside scope", "not covered",
    "not applicable", "please specify", "unable to", "invalid",
    "i need more information", "clarify", "what is your land type",
    "what is your district", "please tell me", "out of scope",
    "not available", "not recognized", "which district", "which land type",
    "not a valid", "soil type", "does not match", "wrong", "not exist",
    "mismatch", "not supported", "not calibrated",
    "कृपया", "जानकारी दें", "जिला बताएं", "भूमि प्रकार", "मुझे बताएं",
    "सही नहीं", "मान्य नहीं", "उपलब्ध नहीं", "please tell", "tell me your", "bata", "batao",
    "kaun sa", "कौन सा"
]


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def has_devanagari(text):
    return bool(re.search(r'[ऀ-ॿ]', str(text)))

def to_latin(text):
    if INDIC_AVAILABLE and has_devanagari(text):
        try:
            return transliterate(str(text), sanscript.DEVANAGARI, sanscript.ITRANS).lower()
        except Exception:
            pass
    return str(text).lower()

def extract_name(kw):
    return re.split(r'\s*\(', kw)[0].strip()

def parse_key_inputs(raw):
    if pd.isna(raw):
        return None, None, None
    s = str(raw).strip()
    if s.lower() in ("no key input", "", "nan"):
        return "NO_KEY", None, None
    s = s.replace("\n", ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]
    subdistrict = land_type = irrigation = None
    for part in parts:
        pl = part.lower()
        if "sub-district" in pl or "subdistrict" in pl:
            subdistrict = re.sub(r"\s*sub-?district\s*", "", pl, flags=re.IGNORECASE).strip()
        elif "no irrigation" in pl:
            irrigation = "No"
        elif "irrigation available" in pl:
            irrigation = "Yes"
        else:
            cleaned = re.sub(r"\s*\bland\b\s*$", "", pl, flags=re.IGNORECASE).strip()
            if cleaned:
                land_type = cleaned
    if land_type:
        land_type = next((c for s, c in LAND_TYPE_SYNONYMS.items() if s in land_type), land_type)
    return subdistrict, land_type, irrigation

def is_monsoon_scenario(scenario_text):
    if pd.isna(scenario_text):
        return False
    sl = str(scenario_text).lower()
    return any(kw in sl for kw in MONSOON_SCENARIO_KEYWORDS)

def find_golden_rows(gd_sheets, district, land_type, irrigation):
    if district not in gd_sheets:
        return pd.DataFrame(), f"Sheet '{district}' not found"
    sheet = gd_sheets[district].copy()
    sheet = sheet[sheet["Crop / Animal"].str.lower().str.contains("rice", na=False)]
    if irrigation == "No":
        sheet = sheet[sheet["Irrigation Available"].isin(["No"])]
    elif irrigation == "Yes":
        sheet = sheet[sheet["Irrigation Available"].isin(["Yes", "Yes/No"])]
    if land_type:
        mask = sheet["Land Type"].str.lower().str.contains(re.escape(land_type), na=False)
        if mask.sum() == 0:
            words = [w for w in land_type.split() if len(w) > 2]
            if words:
                mask = sheet["Land Type"].str.lower().apply(lambda x: any(w in x for w in words))
        sheet = sheet[mask]
    if "Specific Scenario" in sheet.columns:
        sheet = sheet[sheet["Specific Scenario"].apply(is_monsoon_scenario)]
    else:
        print(f"  WARNING: 'Specific Scenario' column missing in '{district}' — skipping scenario filter")
    if sheet.empty:
        return pd.DataFrame(), "No row matches (district + land type + irrigation + monsoon scenario)"
    return sheet, "OK"

def golden_rows_label(golden_rows, district):
    excel_rows = [str(i + 2) for i in golden_rows.index.tolist()]
    return f"{district} · rows {', '.join(excel_rows)}"

def build_union_keywords(golden_rows):
    seen = set()
    union = []
    for _, grow in golden_rows.iterrows():
        AGRONOMIC_COLS = [
            "Expected: Seed Varieties",
            "Expected: Farming Practices",
            "Expected: Fertilizer Dose",
            "Expected: Chemicals",
            "Expected: Infrastructure",
        ]
        kw_str = " ; ".join(str(grow[c]) for c in AGRONOMIC_COLS if c in grow.index and pd.notna(grow[c]))
        if pd.isna(kw_str):
            continue
        for raw in str(kw_str).replace(";", "\n").split("\n"):
            sub_items = [i.strip() for i in re.sub(r'^[•.\s]+', '', raw).strip().split(",") if i.strip()]
            new_kws = [kw for kw in sub_items if kw and kw.lower() not in seen]
            seen.update(kw.lower() for kw in new_kws)
            union.extend(new_kws)
    return union

def split_into_chunks(text):
    raw = re.split(r'[\n]+|(?<=[.!?])\s+', text)
    chunks = []
    for part in raw:
        chunks.extend(re.split(r'\s*[-•*]\s+', part))
    return [c.strip() for c in chunks if len(c.strip()) > 8]


def _rrf_ranks(scores):
    """Return 1-based ranks for a list of scores (highest score → rank 1)."""
    order = sorted(range(len(scores)), key=lambda i: -scores[i])
    ranks = [0] * len(scores)
    for pos, idx in enumerate(order):
        ranks[idx] = pos + 1
    return ranks


def score_keywords(response, keywords):
    """Score keywords against a response using Reciprocal Rank Fusion.

    Pass 1 — exact/name match (fast, always wins).
    Pass 2 — score remaining candidates with fuzzy + semantic.
    Pass 3 — combine signals via RRF; fall back to single-method thresholds
              when only one matcher is available.
    """
    if not keywords or pd.isna(response):
        return None, [], []
    resp = to_latin(str(response))
    n = len(keywords)

    # Pre-compute response chunk embeddings once
    chunk_embs = None
    if SEMANTIC_AVAILABLE:
        chunks = split_into_chunks(resp)
        if chunks:
            chunk_embs = _SEMANTIC_MODEL.encode(chunks, convert_to_tensor=True)

    matched = []
    rrf_indices = []  # indices of keywords that didn't exact/name-match

    # ── Pass 1: exact and name match ─────────────────────────────────────────
    for i, kw in enumerate(keywords):
        kw_lower  = kw.lower()
        name_only = extract_name(kw).lower()
        if kw_lower in resp:
            matched.append(kw)
        elif name_only and len(name_only) > 3 and name_only in resp:
            matched.append(f"{kw} [name match]")
        else:
            rrf_indices.append(i)

    if not rrf_indices:
        pct = round(len(matched) / n * 100, 1)
        return pct, matched, []

    cand_kws = [keywords[i] for i in rrf_indices]

    # ── Pass 2: score each candidate ─────────────────────────────────────────
    fuzzy_scores = []
    if RAPIDFUZZ_AVAILABLE:
        for kw in cand_kws:
            fuzzy_scores.append(fuzz.partial_ratio(kw.lower(), resp) / 100.0)

    semantic_scores = []
    if chunk_embs is not None:
        kw_embs = _SEMANTIC_MODEL.encode(
            [kw.lower() for kw in cand_kws], convert_to_tensor=True
        )
        sims = cos_sim(kw_embs, chunk_embs)  # (n_cands, n_chunks)
        semantic_scores = [row.max().item() for row in sims]

    n_methods = sum([bool(fuzzy_scores), bool(semantic_scores)])

    # ── Pass 3: classify via RRF (or single-method fallback) ─────────────────
    missed = []

    if n_methods == 2:
        fuzzy_ranks    = _rrf_ranks(fuzzy_scores)
        semantic_ranks = _rrf_ranks(semantic_scores)
        for j, kw in enumerate(cand_kws):
            rrf_score = (
                1 / (RRF_K + fuzzy_ranks[j])
                + 1 / (RRF_K + semantic_ranks[j])
            )
            if rrf_score >= RRF_MATCH_THRESHOLD:
                matched.append(f"{kw} [~rrf]")
            else:
                missed.append(kw)

    elif n_methods == 1:
        # Single method: use its raw threshold (preserves old behaviour)
        if fuzzy_scores:
            for j, kw in enumerate(cand_kws):
                if fuzzy_scores[j] >= FUZZY_THRESHOLD / 100.0:
                    matched.append(f"{kw} [~fuzzy]")
                else:
                    missed.append(kw)
        else:
            for j, kw in enumerate(cand_kws):
                if semantic_scores[j] >= SEMANTIC_THRESHOLD:
                    matched.append(f"{kw} [~semantic]")
                else:
                    missed.append(kw)

    else:
        missed.extend(cand_kws)

    pct = round(len(matched) / n * 100, 1)
    return pct, matched, missed


def score_adversarial(response):
    if pd.isna(response) or str(response).strip() == "":
        return "FAIL", "No response recorded"
    resp_lower = str(response).lower()
    if any(phrase in resp_lower for phrase in FLAGGING_PHRASES):
        return "PASS", "Chatbot flagged the issue or asked for clarification"
    return "FAIL", "Chatbot answered directly — did not flag the problem"


# ─── EXCEL STYLE HELPERS ──────────────────────────────────────────────────────

FILL_GREEN     = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW    = PatternFill("solid", fgColor="FFEB9C")
FILL_RED       = PatternFill("solid", fgColor="FFC7CE")
FILL_GREY      = PatternFill("solid", fgColor="D9D9D9")
FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")
FILL_AUDIT_HDR = PatternFill("solid", fgColor="2E75B6")

def style_header(ws, audit_col_idx=None):
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = FILL_AUDIT_HDR if (audit_col_idx and i == audit_col_idx) else FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def wrap_data_cells(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading files...")
    pb = pd.read_excel(PROMPT_BANK_FILE)
    pb.columns = [c.strip() for c in pb.columns]
    gd_raw = pd.read_excel(GOLDEN_DATASET_FILE, sheet_name=None)
    gd = {name: df.rename(columns=lambda c: str(c).strip()) for name, df in gd_raw.items()}
    pb = pb.dropna(subset=["Q#"]).reset_index(drop=True)
    print(f"  Prompt rows to score: {len(pb)}")

    std_results = []

    for _, row in pb.iterrows():
        sr         = int(row["Q#"])
        prompt     = str(row["English Query"])
        response   = row["English response (normal onset)"]
        district   = str(row["District"]).strip() if pd.notna(row["District"]) else ""
        local_term = str(row["Land Type (local term)"]).strip().lower() if pd.notna(row["Land Type (local term)"]) else ""
        land_type  = LAND_TYPE_SYNONYMS.get(local_term, local_term) if local_term else None
        irrigation = "Yes" if str(row["Irrigation"]).strip() == "Yes" else "No"

        rec = {
            "Q#"                  : sr,
            "Prompt"              : prompt,
            "Chatbot Response"    : str(response) if pd.notna(response) else "",
            "Key Inputs"          : "",
            "Keywords (Reference)": "",
            "Keywords Captured"   : "",
            "Keywords Missed"     : "",
            "Score (%)"           : "",
            "Golden Dataset Rows" : "",
            "Notes"               : "",
        }

        if not district:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "District not specified"
            std_results.append(rec)
            continue

        rec["Key Inputs"] = f"{district} / {land_type or 'land type not specified'} / {'No irrigation' if irrigation == 'No' else 'Irrigation available'}"

        golden_rows, status = find_golden_rows(gd, district, land_type, irrigation)
        if golden_rows.empty:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = f"No golden dataset match — {status}"
            std_results.append(rec)
            continue

        missing_input = "not specified" in rec["Key Inputs"].lower()
        asks_clarification = "?" in str(response) and any(
            p in str(response).lower() for p in [
                "land type", "land-type", "upland", "lowland", "midland",
                "district", "block", "tehsil", "subdistrict",
                "irrigation", "soil type", "zameen", "bhoomi",
                "जमीन", "जिला", "सिंचाई", "भूमि"
            ]
        )
        if missing_input and asks_clarification:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "Correct behaviour — chatbot asked for missing key input"
            std_results.append(rec)
            continue

        rec["Golden Dataset Rows"] = golden_rows_label(golden_rows, district)
        union_kw = build_union_keywords(golden_rows)
        rec["Keywords (Reference)"] = ("• " + "\n• ".join(union_kw)) if union_kw else ""

        pct, matched, missed = score_keywords(response, union_kw)
        if pct is None:
            rec["Score (%)"] = "N/A"
            rec["Notes"] = "Keywords column empty in golden dataset"
        else:
            rec["Score (%)"]         = pct
            rec["Keywords Captured"] = ("• " + "\n• ".join(matched)) if matched else ""
            rec["Keywords Missed"]   = ("• " + "\n• ".join(missed)) if missed else ""
            rec["Notes"] = "Low score — review chatbot response" if pct < 30 else ("Partial — some key advice missing" if pct < 60 else "")
        std_results.append(rec)

    std_df = pd.DataFrame(std_results)
    scores  = pd.to_numeric(std_df["Score (%)"], errors="coerce").dropna()

    summary_df = pd.DataFrame({
        "Metric": [
            "Total prompts tested",
            "Standard prompts scored", "Standard prompts – no match / no key input",
            "Average keyword score (%)", "Score >= 60% (good)", "Score 30-59% (partial)",
            "Score < 30% (poor)",
        ],
        "Value": [
            len(std_df),
            len(scores), len(std_df) - len(scores),
            f"{scores.mean():.1f}%" if len(scores) else "–",
            int((scores >= 60).sum()) if len(scores) else 0,
            int(((scores >= 30) & (scores < 60)).sum()) if len(scores) else 0,
            int((scores < 30).sum()) if len(scores) else 0,
        ],
    })

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]
        style_header(ws)
        set_col_widths(ws, {"A": 50, "B": 20})
        wrap_data_cells(ws)

        std_df.to_excel(writer, index=False, sheet_name="Standard Scores")
        ws = writer.sheets["Standard Scores"]
        style_header(ws, audit_col_idx=9)
        set_col_widths(ws, {
            "A": 6,  "B": 50, "C": 55, "D": 35,
            "E": 55, "F": 55, "G": 55, "H": 10,
            "I": 28, "J": 40,
        })
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            score_cell = row[7]
            val = score_cell.value
            if val == "N/A":
                score_cell.fill = FILL_GREY
            else:
                try:
                    s = float(val)
                    score_cell.fill = FILL_GREEN if s >= 60 else (FILL_YELLOW if s >= 30 else FILL_RED)
                except (TypeError, ValueError):
                    pass
        wrap_data_cells(ws)

    print(f"\n{'='*50}\nRESULTS SUMMARY\n{'='*50}")
    print(f"Total : {len(std_df)} prompts, {len(scores)} scored")
    if len(scores):
        print(f"  Avg score : {scores.mean():.1f}%  |  >=60%: {int((scores>=60).sum())}  |  30-59%: {int(((scores>=30)&(scores<60)).sum())}  |  <30%: {int((scores<30).sum())}")
    print(f"\nOutput saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
