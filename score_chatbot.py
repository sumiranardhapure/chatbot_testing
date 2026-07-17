#!/usr/bin/env python3
"""
CRIDA DACP Chatbot Scoring Script
===================================
Scores chatbot responses (col H) against golden dataset keywords (col R).

Setup (run once in terminal):
    pip3 install pandas openpyxl rapidfuzz indic-transliteration

Usage:
    1. Place this script, prompt bank, and golden dataset in the same folder.
    2. Update FILE PATHS below if your filenames differ.
    3. Run: python3 score_chatbot.py

Output:
    Chatbot_Scoring_Results.xlsx — three sheets: Summary, Standard Scores, Adversarial Scores.
"""

import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

try:
    from rapidfuzz import fuzz
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False
    print("WARNING: rapidfuzz not installed. Run: pip3 install rapidfuzz")

try:
    from indic_transliteration import sanscript
    from indic_transliteration.sanscript import transliterate
    INDIC_AVAILABLE = True
except ImportError:
    INDIC_AVAILABLE = False


# ─── FILE PATHS ───────────────────────────────────────────────────────────────
PROMPT_BANK_FILE    = "prompt_file_v2.xlsx"
GOLDEN_DATASET_FILE = "golden_dataset_v2.xlsx"
OUTPUT_FILE         = "Chatbot_Scoring_Results_v4.xlsx"


# ─── SUB-DISTRICT → DISTRICT MAPPING ─────────────────────────────────────────
SUBDISTRICT_TO_DISTRICT = {
    "sukma":          "Sukma",
    "bijapur":        "Bijapur",
    "bhopalpatnam":   "Bijapur",
    "dantewada":      "Dantewada",
    "geedam":         "Dantewada",
    "kuakonda":       "Dantewada",
    "kondagaon":      "Kondagaon",
    "odgil":          "Kondagaon",
    "koriya":         "Koriya",
    "sonhat":         "Koriya",
    "jashpur":        "Jashpur",
    "bagicha":        "Jashpur",
    "pathalgaon":     "Jashpur",
    "pharagaon":      "Jashpur",
    "surguja":        "Surguja",
    "ambikapur":      "Surguja",
    "lundra":         "Surguja",
    "lakhanpur":      "Surguja",
    "surajpur":       "Surajpur",
    "chhindgarh":     "Sukma",
    "sitapur":        "Surajpur",
    "balrampur":      "Balrampur",
    "ramanujnagar":   "Balrampur",
    "batouli":        "Balrampur",
    "narayanpur":     "Narayanpur",
    "orchaa":         "Narayanpur",
    "kanker":         "Kanker",
    "charma":         "Kanker",
    "bhanupratappur": "Kanker",
    "pakhanjur":      "Kanker",
}
KNOWN_DISTRICTS = list(set(SUBDISTRICT_TO_DISTRICT.values()))

def extract_explicit_districts(text):
    if not text:
        return None
    text_lower = str(text).lower()
    for d in KNOWN_DISTRICTS:
        if d.lower() in text_lower:
            return d
    return None

# ─── LAND TYPE SYNONYM MAP ────────────────────────────────────────────────────
LAND_TYPE_SYNONYMS = {
    "marhan":  "upland",
    "tikra":   "upland",
    "gabhar":  "lowland",
    "mal":     "midland",
    "bharri":  "upland",
    "bhata":   "upland",
    "kanhar":  "lowland",
    "matasi":  "midland",
    "upariya": "upland",
    "dorsa":   "midland",
}


# ─── MONSOON-RELEVANT SCENARIO FILTER ────────────────────────────────────────
MONSOON_SCENARIO_KEYWORDS = [
    "delayed",
    "onset",
    "early season drought",
    "normal",
]


# ─── ADVERSARIAL FLAGGING PHRASES ────────────────────────────────────────────
FLAGGING_PHRASES = [
    "does not exist", "not valid", "incorrect", "not found", "please clarify",
    "please provide", "cannot find", "outside scope", "not covered",
    "not applicable", "please specify", "unable to", "invalid",
    "i need more information", "clarify", "what is your land type",
    "what is your district", "please tell me", "out of scope",
    "not available", "not recognized", "which district", "which land type",
    "not a valid", "soil type", "does not match", "wrong", "not exist",
    "mismatch", "not supported", "not calibrated",
    "कृपया", "जानकारी दें", "जिला बताएं", "भूमि प्रकार", "मुझे बताएं",
    "सही नहीं", "मान्य नहीं", "उपलब्ध नहीं", "please tell", "tell me your", "bata", "batao"
    "kaun sa", "कौन सा"
]


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def has_devanagari(text):
    return bool(re.search(r'[ऀ-ॿ]', str(text)))

def to_latin(text):
    if INDIC_AVAILABLE and has_devanagari(text):
        try:
            return transliterate(str(text), sanscript.DEVANAGARI, sanscript.ITRANS).lower()
        except Exception:
            pass
    return str(text).lower()

def extract_name(kw):
    return re.split(r'\s*\(', kw)[0].strip()

def parse_key_inputs(raw):
    if pd.isna(raw):
        return None, None, None
    s = str(raw).strip()
    if s.lower() in ("no key input", "", "nan"):
        return "NO_KEY", None, None
    s = s.replace("\n", ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]
    subdistrict = land_type = irrigation = None
    for part in parts:
        pl = part.lower()
        if "sub-district" in pl or "subdistrict" in pl:
            subdistrict = re.sub(r"\s*sub-?district\s*", "", pl, flags=re.IGNORECASE).strip()
        elif "no irrigation" in pl:
            irrigation = "No"
        elif "irrigation available" in pl:
            irrigation = "Yes"
        else:
            cleaned = re.sub(r"\s*\bland\b\s*$", "", pl, flags=re.IGNORECASE).strip()
            if cleaned:
                land_type = cleaned
    if land_type:
        land_type = next((c for s, c in LAND_TYPE_SYNONYMS.items() if s in land_type), land_type)
    return subdistrict, land_type, irrigation

def is_monsoon_scenario(scenario_text):
    if pd.isna(scenario_text):
        return False
    sl = str(scenario_text).lower()
    return any(kw in sl for kw in MONSOON_SCENARIO_KEYWORDS)

def find_golden_rows(gd_sheets, district, land_type, irrigation):
    if district not in gd_sheets:
        return pd.DataFrame(), f"Sheet '{district}' not found"
    sheet = gd_sheets[district].copy()
    sheet = sheet[sheet["Crop / Animal"].str.lower().str.contains("rice", na=False)]
    if irrigation == "No":
        sheet = sheet[sheet["Irrigation Available"].isin(["No"])]
    elif irrigation == "Yes":
        sheet = sheet[sheet["Irrigation Available"].isin(["Yes", "Yes/No"])]
    if land_type:
        mask = sheet["Land Type"].str.lower().str.contains(re.escape(land_type), na=False)
        if mask.sum() == 0:
            words = [w for w in land_type.split() if len(w) > 2]
            if words:
                mask = sheet["Land Type"].str.lower().apply(lambda x: any(w in x for w in words))
        sheet = sheet[mask]
    if "Specific Scenario" in sheet.columns:
        sheet = sheet[sheet["Specific Scenario"].apply(is_monsoon_scenario)]
    else:
        print(f"  WARNING: 'Specific Scenario' column missing in '{district}' — skipping scenario filter")
    if sheet.empty:
        return pd.DataFrame(), "No row matches (district + land type + irrigation + monsoon scenario)"
    return sheet, "OK"

def golden_rows_label(golden_rows, district):
    excel_rows = [str(i + 2) for i in golden_rows.index.tolist()]
    return f"{district} · rows {', '.join(excel_rows)}"

def build_union_keywords(golden_rows):
    seen = set()
    union = []
    for _, grow in golden_rows.iterrows():
        AGRONOMIC_COLS = [
            "Expected: Seed Varieties",
            "Expected: Farming Practices",
            "Expected: Fertilizer Dose",
            "Expected: Chemicals",
            "Expected: Infrastructure",
        ]
        kw_str = " ; ".join(str(grow[c]) for c in AGRONOMIC_COLS if c in grow.index and pd.notna(grow[c]))
        if pd.isna(kw_str):
            continue
        for raw in str(kw_str).replace(";", "\n").split("\n"):
            sub_items = [i.strip() for i in raw.lstrip(".").strip().split(",") if i.strip()]
            new_kws = [kw for kw in sub_items if kw and kw.lower() not in seen]
            seen.update(kw.lower() for kw in new_kws)
            union.extend(new_kws)
    return union

def score_keywords(response, keywords, fuzzy_threshold=75):
    if not keywords or pd.isna(response):
        return None, [], []
    resp = to_latin(str(response))
    matched, missed = [], []
    for kw in keywords:
        kw_lower  = kw.lower()
        name_only = extract_name(kw).lower()
        if kw_lower in resp:
            matched.append(kw)
        elif name_only and len(name_only) > 3 and name_only in resp:
            matched.append(f"{kw} [name match]")
        elif RAPIDFUZZ_AVAILABLE and fuzz.partial_ratio(kw_lower, resp) >= fuzzy_threshold:
            matched.append(f"{kw} [~fuzzy]")
        else:
            missed.append(kw)
    pct = round(len(matched) / len(keywords) * 100, 1) if keywords else None
    return pct, matched, missed

def score_adversarial(response):
    if pd.isna(response) or str(response).strip() == "":
        return "FAIL", "No response recorded"
    resp_lower = str(response).lower()
    if any(phrase in resp_lower for phrase in FLAGGING_PHRASES):
        return "PASS", "Chatbot flagged the issue or asked for clarification"
    return "FAIL", "Chatbot answered directly — did not flag the problem"


# ─── EXCEL STYLE HELPERS ──────────────────────────────────────────────────────

FILL_GREEN     = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW    = PatternFill("solid", fgColor="FFEB9C")
FILL_RED       = PatternFill("solid", fgColor="FFC7CE")
FILL_GREY      = PatternFill("solid", fgColor="D9D9D9")
FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")
FILL_AUDIT_HDR = PatternFill("solid", fgColor="2E75B6")

def style_header(ws, audit_col_idx=None):
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = FILL_AUDIT_HDR if (audit_col_idx and i == audit_col_idx) else FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def wrap_data_cells(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading files...")
    pb = pd.read_excel(PROMPT_BANK_FILE)
    pb.columns = [c.strip() for c in pb.columns]
    gd_raw = pd.read_excel(GOLDEN_DATASET_FILE, sheet_name=None)
    gd = {name: df.rename(columns=lambda c: str(c).strip()) for name, df in gd_raw.items()}
    pb = pb[pb["Prompt Type"].isin(["Standard prompt", "Adversarial Prompt"])]
    pb = pb.dropna(subset=["Sr. No."]).reset_index(drop=True)
    print(f"  Prompt rows to score: {len(pb)}")

    std_results, adv_results, unmapped = [], [], set()

    for _, row in pb.iterrows():
        sr         = int(row["Sr. No."])
        ptype      = row["Prompt Type"]
        prompt     = str(row["Prompt Questions"])
        response   = row["Chatbot response"]
        key_inputs = row["Key Inputs"]
        flags      = row["Flags for incorrect prompts"]

        if ptype == "Standard prompt":
            rec = {
                "Sr. No."             : sr,
                "Prompt"              : prompt,
                "Chatbot Response"    : str(response) if pd.notna(response) else "",
                "Key Inputs"          : "",
                "Keywords (Reference)": "",
                "Keywords Captured"   : "",
                "Keywords Missed"     : "",
                "Score (%)"           : "",
                "Golden Dataset Rows" : "",
                "Notes"               : "",
            }

            subdistrict, land_type, irrigation = parse_key_inputs(key_inputs)
            if subdistrict in (None, "NO_KEY"):
                rec["Score (%)"] = "N/A"
                rec["Notes"] = "No key input in prompt — unevaluable"
                std_results.append(rec); continue

            explicit = extract_explicit_districts(key_inputs)
            district = explicit or SUBDISTRICT_TO_DISTRICT.get(subdistrict.lower() if subdistrict else "","")
            if not district:
                rec["Score (%)"] = "N/A"
                rec["Notes"] = f"'{subdistrict}' not in sub-district map"
                unmapped.add(subdistrict)
                std_results.append(rec); continue

            rec["Key Inputs"] = f"{district} / {land_type or 'land type not specified'} / {'No irrigation' if irrigation == 'No' else 'Irrigation available' if irrigation == 'Yes' else 'irrigation not specified'}"

            golden_rows, status = find_golden_rows(gd, district, land_type, irrigation)
            if golden_rows.empty:
                rec["Score (%)"] = "N/A"
                rec["Notes"] = f"No golden dataset match — {status}"
                std_results.append(rec); continue
            
            missing_input = "not specified" in rec["Key Inputs"].lower()
            asks_clarification = "?" in str(response) and any(
                p in str(response).lower() for p in [
                    "land type", "land-type", "upland", "lowland", "midland",
                    "district", "block", "tehsil", "subdistrict",
                    "irrigation", "soil type", "zameen", "bhoomi",
                    "जमीन", "जिला", "सिंचाई", "भूमि"
                ]
            )
            if missing_input and asks_clarification:
                rec["Score (%)"] = "N/A"
                rec["Notes"] = "Correct behaviour — chatbot asked for missing key input"
                std_results.append(rec)
                continue

            rec["Golden Dataset Rows"] = golden_rows_label(golden_rows, district)
            union_kw = build_union_keywords(golden_rows)
            rec["Keywords (Reference)"] = ("• " + "\n• ".join(union_kw)) if union_kw else ""

            pct, matched, missed = score_keywords(response, union_kw)
            if pct is None:
                rec["Score (%)"] = "N/A"
                rec["Notes"] = "Keywords column empty in golden dataset"
            else:
                rec["Score (%)"]         = pct
                rec["Keywords Captured"] = ("• " + "\n• ".join(matched)) if matched else ""
                rec["Keywords Missed"] = ("• " + "\n• ".join(missed)) if missed else ""
                rec["Notes"] = "Low score — review chatbot response" if pct < 30 else ("Partial — some key advice missing" if pct < 60 else "")
            std_results.append(rec)

        elif ptype == "Adversarial Prompt":
            adv_result, adv_note = score_adversarial(response)
            adv_results.append({
                "Sr. No."          : sr,
                "Prompt"           : prompt,
                "Chatbot Response" : str(response) if pd.notna(response) else "",
                "Expected Behavior": str(flags) if pd.notna(flags) else "",
                "Result"           : adv_result,
                "Notes"            : adv_note,
            })

    std_df     = pd.DataFrame(std_results)
    adv_df     = pd.DataFrame(adv_results)
    scores     = pd.to_numeric(std_df["Score (%)"], errors="coerce").dropna()
    adv_scored = adv_df["Result"] if len(adv_df) else pd.Series(dtype=str)
    
    correct_clarification = int(std_df["Notes"].str.contains("Correct behaviour", na=False).sum())
    summary_df = pd.DataFrame({
        "Metric": [
            "Total prompts tested", "– Standard prompts", "– Adversarial prompts", "",
            "Standard prompts scored", "Standard prompts – no match / no key input",
            "Average keyword score (%)", "Score >= 60% (good)", "Score 30-59% (partial)",
            "Score < 30% (poor)", "",
            "Correctly handled missing input (N/A)",
            "Adversarial: PASS (correctly flagged issue)",
            "Adversarial: FAIL (answered directly without flagging)",
        ],
        "Value": [
            len(std_df) + len(adv_df), len(std_df), len(adv_df), "",
            len(scores), len(std_df) - len(scores),
            f"{scores.mean():.1f}%" if len(scores) else "–",
            int((scores >= 60).sum()) if len(scores) else 0,
            int(((scores >= 30) & (scores < 60)).sum()) if len(scores) else 0,
            int((scores < 30).sum()) if len(scores) else 0, "",
            correct_clarification,
            int((adv_scored == "PASS").sum()) if len(adv_scored) else 0,
            int((adv_scored == "FAIL").sum()) if len(adv_scored) else 0,
        ],
    })

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]
        style_header(ws)
        set_col_widths(ws, {"A": 50, "B": 20})
        wrap_data_cells(ws)

        # A: Sr.No  B: Prompt  C: Response  D: Key Inputs
        # E: Keywords(Reference)  F: Keywords Captured  G: Keywords Missed
        # H: Score(%)  I: Golden Dataset Rows  J: Notes
        std_df.to_excel(writer, index=False, sheet_name="Standard Scores")
        ws = writer.sheets["Standard Scores"]
        style_header(ws, audit_col_idx=9)
        set_col_widths(ws, {
            "A": 6,  "B": 50, "C": 55, "D": 35,
            "E": 55, "F": 55, "G": 55, "H": 10,
            "I": 28, "J": 40,
        })
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            score_cell = row[7]
            val = score_cell.value
            if val == "N/A":
                score_cell.fill = FILL_GREY
            else:
                try:
                    s = float(val)
                    score_cell.fill = FILL_GREEN if s >= 60 else (FILL_YELLOW if s >= 30 else FILL_RED)
                except (TypeError, ValueError):
                    pass
        wrap_data_cells(ws)

        adv_df.to_excel(writer, index=False, sheet_name="Adversarial Scores")
        ws = writer.sheets["Adversarial Scores"]
        style_header(ws)
        set_col_widths(ws, {"A": 6, "B": 55, "C": 60, "D": 60, "E": 10, "F": 50})
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[4].fill = FILL_GREEN if row[4].value == "PASS" else (FILL_RED if row[4].value == "FAIL" else row[4].fill)
        wrap_data_cells(ws)

    print(f"\n{'='*50}\nRESULTS SUMMARY\n{'='*50}")
    print(f"Standard : {len(std_df)} total, {len(scores)} scored")
    correct_clarification = std_df["Notes"].str.contains("Correct behaviour", na=False).sum()
    print(f"  Correctly asked for missing key input: {correct_clarification} rows")
    if len(scores):
        print(f"  Avg score : {scores.mean():.1f}%  |  >=60%: {int((scores>=60).sum())}  |  30-59%: {int(((scores>=30)&(scores<60)).sum())}  |  <30%: {int((scores<30).sum())}")
    print(f"Adversarial : PASS {int((adv_scored=='PASS').sum())} / FAIL {int((adv_scored=='FAIL').sum())}")
    if unmapped:
        print(f"\nUnmapped sub-districts: {', '.join(sorted(unmapped))}")
    print(f"\nOutput saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()